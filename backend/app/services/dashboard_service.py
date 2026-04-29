import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import desc, distinct, func, nulls_last, or_
from sqlalchemy.orm import Session

from app.models import CrawlEdge, CrawlRun, Product, ProductSnapshot
from app.models.user import User


def is_admin(user: User) -> bool:
    return user.role == "admin"


def decimal_to_text(value) -> str:
    if value is None:
        return ""
    return format(value, "f").rstrip("0").rstrip(".")


def get_dashboard_overview(db: Session, current_user: User) -> dict:
    product_q = db.query(Product)
    snapshot_q = db.query(ProductSnapshot)
    run_q = db.query(CrawlRun)
    edge_q = db.query(CrawlEdge)
    if not is_admin(current_user):
        product_q = product_q.filter(Product.user_id == current_user.id)
        snapshot_q = snapshot_q.filter(ProductSnapshot.user_id == current_user.id)
        run_q = run_q.filter(CrawlRun.user_id == current_user.id)
        edge_q = edge_q.filter(CrawlEdge.user_id == current_user.id)

    total_products = product_q.with_entities(func.count(Product.id)).scalar() or 0
    total_snapshots = snapshot_q.with_entities(func.count(ProductSnapshot.id)).scalar() or 0
    total_runs = run_q.with_entities(func.count(CrawlRun.id)).scalar() or 0
    completed_runs = (
        run_q.with_entities(func.count(CrawlRun.id))
        .filter(CrawlRun.status == "completed")
        .scalar()
        or 0
    )
    total_edges = edge_q.with_entities(func.count(CrawlEdge.id)).scalar() or 0
    related_sources = (
        snapshot_q.with_entities(func.count(ProductSnapshot.id))
        .filter(ProductSnapshot.source == "related")
        .scalar()
        or 0
    )
    distinct_goods = snapshot_q.with_entities(func.count(distinct(ProductSnapshot.goods_id))).scalar() or 0
    latest_snapshot = snapshot_q.with_entities(func.max(ProductSnapshot.scraped_at)).scalar()

    return {
        "total_products": total_products,
        "total_snapshots": total_snapshots,
        "total_runs": total_runs,
        "completed_runs": completed_runs,
        "total_edges": total_edges,
        "related_sources": related_sources,
        "distinct_snapshot_goods": distinct_goods,
        "latest_snapshot_at": latest_snapshot.isoformat() if latest_snapshot else None,
    }


def get_dashboard_runs(db: Session, current_user: User, limit: int = 20) -> list[dict]:
    query = db.query(CrawlRun)
    snapshot_count_query = db.query(
        ProductSnapshot.run_uuid,
        func.count(distinct(ProductSnapshot.goods_id)).label("collected_count"),
    )
    if not is_admin(current_user):
        query = query.filter(CrawlRun.user_id == current_user.id)
        snapshot_count_query = snapshot_count_query.filter(ProductSnapshot.user_id == current_user.id)
    runs = query.order_by(desc(CrawlRun.started_at)).limit(limit).all()
    run_uuids = [item.run_uuid for item in runs]
    snapshot_counts = {}
    if run_uuids:
        snapshot_counts = {
            run_uuid: count
            for run_uuid, count in (
                snapshot_count_query
                .filter(ProductSnapshot.run_uuid.in_(run_uuids))
                .group_by(ProductSnapshot.run_uuid)
                .all()
            )
        }
    return [
        {
            "run_uuid": item.run_uuid,
            "status": item.status,
            "started_at": item.started_at.isoformat() if item.started_at else None,
            "ended_at": item.ended_at.isoformat() if item.ended_at else None,
            "total_collected": snapshot_counts.get(item.run_uuid, item.total_collected),
            "notes": item.notes or "",
        }
        for item in runs
    ]


def get_dashboard_sources(db: Session, current_user: User) -> list[dict]:
    query = db.query(ProductSnapshot.source, func.count(ProductSnapshot.id))
    if not is_admin(current_user):
        query = query.filter(ProductSnapshot.user_id == current_user.id)
    rows = (
        query
        .group_by(ProductSnapshot.source)
        .order_by(desc(func.count(ProductSnapshot.id)))
        .all()
    )
    return [
        {
            "source": source or "(empty)",
            "count": count,
        }
        for source, count in rows
    ]


def get_dashboard_edges(db: Session, current_user: User, limit: int = 20) -> list[dict]:
    query = db.query(
            CrawlEdge.from_goods_id,
            CrawlEdge.to_goods_id,
            CrawlEdge.relation_type,
            func.count(CrawlEdge.id).label("count"),
        )
    if not is_admin(current_user):
        query = query.filter(CrawlEdge.user_id == current_user.id)
    rows = (
        query
        .group_by(CrawlEdge.from_goods_id, CrawlEdge.to_goods_id, CrawlEdge.relation_type)
        .order_by(desc(func.count(CrawlEdge.id)))
        .limit(limit)
        .all()
    )
    return [
        {
            "from_goods_id": from_goods_id,
            "to_goods_id": to_goods_id,
            "relation_type": relation_type,
            "count": count,
        }
        for from_goods_id, to_goods_id, relation_type, count in rows
    ]


def get_dashboard_products(
    db: Session,
    current_user: User,
    keyword: str = "",
    page: int = 1,
    page_size: int = 30,
    sort_by: str = "last_seen_at",
    sort_order: str = "desc",
) -> dict:
    query = db.query(Product)
    if not is_admin(current_user):
        query = query.filter(Product.user_id == current_user.id)
    if keyword.strip():
        pattern = f"%{keyword.strip()}%"
        query = query.filter(
            or_(
                Product.goods_id.ilike(pattern),
                Product.current_title.ilike(pattern),
                Product.current_full_title.ilike(pattern),
            )
        )

    total = query.count()

    sort_mapping = {
        "goods_id": Product.goods_id,
        "title": Product.current_title,
        "price_text": Product.current_price_text,
        "sales_text": Product.current_sales_text,
        "sales_value": Product.current_sales_value,
        "star_rating": Product.current_star_rating,
        "review_count": Product.current_review_count,
        "listing_time": Product.current_listing_time,
        "raw_text": Product.current_raw_text,
        "raw_html": Product.current_raw_html,
        "last_source": Product.last_source,
        "last_seen_at": Product.last_seen_at,
        "updated_at": Product.updated_at,
    }
    sort_column = sort_mapping.get(sort_by, Product.last_seen_at)
    order_fn = desc if sort_order == "desc" else lambda column: column.asc()
    sort_expr = order_fn(sort_column)
    if sort_by == "sales_value":
        sort_expr = nulls_last(sort_expr)
    rows = (
        query.order_by(sort_expr, desc(Product.updated_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    items = [
        {
            "goods_id": item.goods_id,
            "title": item.current_title or "",
            "full_title": item.current_full_title or "",
            "price_text": item.current_price_text or "",
            "sales_text": item.current_sales_text or "",
            "sales_value": item.current_sales_value,
            "star_rating": item.current_star_rating or "",
            "review_count": item.current_review_count,
            "listing_time": item.current_listing_time or "",
            "raw_text": item.current_raw_text or "",
            "raw_html": item.current_raw_html or "",
            "listing_length_cm": decimal_to_text(item.listing_length_cm),
            "listing_width_cm": decimal_to_text(item.listing_width_cm),
            "listing_height_cm": decimal_to_text(item.listing_height_cm),
            "listing_weight_g": decimal_to_text(item.listing_weight_g),
            "listing_declared_price": decimal_to_text(item.listing_declared_price),
            "listing_suggested_price": decimal_to_text(item.listing_suggested_price),
            "last_source": item.last_source or "",
            "last_seen_at": item.last_seen_at.isoformat() if item.last_seen_at else None,
            "updated_at": item.updated_at.isoformat() if item.updated_at else None,
        }
        for item in rows
    ]
    return {
        "items": items,
        "page": page,
        "page_size": page_size,
        "total": total,
        "total_pages": (total + page_size - 1) // page_size if page_size else 0,
        "sort_by": sort_by,
        "sort_order": sort_order,
        "keyword": keyword,
    }


def export_products_excel(
    db: Session,
    current_user: User,
    keyword: str = "",
    sort_by: str = "last_seen_at",
    sort_order: str = "desc",
) -> bytes:
    """查询全量（无分页）商品并生成 Excel 文件，返回二进制内容。"""
    query = db.query(Product)
    if not is_admin(current_user):
        query = query.filter(Product.user_id == current_user.id)
    if keyword.strip():
        pattern = f"%{keyword.strip()}%"
        query = query.filter(
            or_(
                Product.goods_id.ilike(pattern),
                Product.current_title.ilike(pattern),
                Product.current_full_title.ilike(pattern),
            )
        )

    sort_mapping = {
        "goods_id": Product.goods_id,
        "title": Product.current_title,
        "price_text": Product.current_price_text,
        "sales_text": Product.current_sales_text,
        "sales_value": Product.current_sales_value,
        "star_rating": Product.current_star_rating,
        "review_count": Product.current_review_count,
        "listing_time": Product.current_listing_time,
        "raw_text": Product.current_raw_text,
        "last_source": Product.last_source,
        "last_seen_at": Product.last_seen_at,
        "updated_at": Product.updated_at,
    }
    sort_column = sort_mapping.get(sort_by, Product.last_seen_at)
    order_fn = desc if sort_order == "desc" else lambda col: col.asc()
    sort_expr = order_fn(sort_column)
    if sort_by == "sales_value":
        sort_expr = nulls_last(sort_expr)
    rows = query.order_by(sort_expr).all()

    # ── 构建 Workbook ─────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "商品数据"

    headers = [
        "商品 ID", "标题", "完整标题", "价格", "销量文本", "销量值", "星级", "评价数",
        "上架时间", "最长边(cm)", "次长边(cm)", "最短边(cm)", "重量(g)",
        "申报价", "建议零售价", "来源", "最后出现", "更新时间",
    ]

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列宽
    col_widths = [20, 36, 50, 12, 14, 12, 10, 10, 20, 14, 14, 14, 12, 12, 14, 14, 20, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append([
            row.goods_id or "",
            row.current_title or "",
            row.current_full_title or "",
            row.current_price_text or "",
            row.current_sales_text or "",
            row.current_sales_value,
            row.current_star_rating or "",
            row.current_review_count,
            row.current_listing_time or "",
            decimal_to_text(row.listing_length_cm),
            decimal_to_text(row.listing_width_cm),
            decimal_to_text(row.listing_height_cm),
            decimal_to_text(row.listing_weight_g),
            decimal_to_text(row.listing_declared_price),
            decimal_to_text(row.listing_suggested_price),
            row.last_source or "",
            row.last_seen_at.replace(tzinfo=None) if row.last_seen_at else None,
            row.updated_at.replace(tzinfo=None) if row.updated_at else None,
        ])

    # 日期列格式
    date_fmt = "YYYY-MM-DD HH:MM:SS"
    for row_idx in range(2, len(rows) + 2):
        ws.cell(row=row_idx, column=17).number_format = date_fmt
        ws.cell(row=row_idx, column=18).number_format = date_fmt

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
